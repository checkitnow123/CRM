"""Export client list to Excel (.xlsx)."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from logic.alerts import effective_alert_days
from logic.clients import (
    client_matches_query,
    is_client_closed,
    load_clients,
    normalize_client_numbers,
)
from logic.constants import MODE_MILESTONE
from logic.dates import days_remaining, format_end_date
from logic.extra_i18n import extra_t
from logic.i18n import mode_label, t
from logic.milestones import milestone_all_complete, milestone_done_count, next_scheduled_visit


def _filter_clients(
    clients: list[dict[str, Any]],
    q: str = "",
    group: str = "",
    account: str = "all",
) -> list[dict[str, Any]]:
    out = sorted(clients, key=lambda c: c["client_name"].lower())
    if account == "active":
        out = [c for c in out if not is_client_closed(c)]
    elif account == "closed":
        out = [c for c in out if is_client_closed(c)]
    if group == "__none__":
        out = [c for c in out if not (c.get("group") or "").strip()]
    elif group:
        out = [c for c in out if (c.get("group") or "").strip() == group]
    if q.strip():
        out = [c for c in out if client_matches_query(c, q)]
    return out


def _export_headers(L: str) -> list[str]:
    return [
        t(L, "col_client"),
        extra_t(L, "col_group"),
        extra_t(L, "col_client_no"),
        extra_t(L, "col_contact"),
        extra_t(L, "col_email"),
        extra_t(L, "export_col_service"),
        extra_t(L, "col_mode"),
        extra_t(L, "export_col_account"),
        extra_t(L, "col_end"),
        extra_t(L, "export_col_start"),
        extra_t(L, "export_col_days"),
        extra_t(L, "col_status"),
        extra_t(L, "col_next"),
        extra_t(L, "export_col_milestone"),
        extra_t(L, "export_col_last_reminder"),
    ]


def _export_row(client: dict[str, Any], L: str) -> list[Any]:
    closed = is_client_closed(client)
    days = effective_alert_days(client) if client["mode"] == MODE_MILESTONE else days_remaining(client["end_date"])
    is_ms = client["mode"] == MODE_MILESTONE
    done = milestone_done_count(client) if is_ms else 0
    total = int(client.get("total_milestones") or 0)
    if is_ms and milestone_all_complete(client):
        ms_text = t(L, "all_sessions_done")
    elif is_ms and total:
        ms_text = t(L, "progress", done=done, total=total)
    else:
        ms_text = "—"
    nxt = next_scheduled_visit(client) if is_ms and not milestone_all_complete(client) else None
    next_date = nxt[1] if nxt else None
    next_text = format_end_date(next_date) if next_date else "—"
    if closed:
        status = extra_t(L, "status_closed")
        account = extra_t(L, "clients_account_closed")
    else:
        from logic.dates import status_badge_label, status_badge_tier

        status = status_badge_label(L, status_badge_tier(days))
        account = extra_t(L, "clients_account_active")
    last = client.get("last_reminder_at") or ""
    if isinstance(last, str) and len(last) >= 10:
        last = last[:10]
    elif not last:
        last = "—"
    numbers = normalize_client_numbers(client.get("client_numbers"))
    return [
        client.get("client_name") or "",
        client.get("group") or "",
        ", ".join(numbers),
        client.get("contact_name") or "",
        client.get("client_email") or "",
        client.get("service_label") or "",
        mode_label(L, client["mode"]),
        account,
        format_end_date(client["end_date"]),
        format_end_date(client.get("start_date") or ""),
        days,
        status,
        next_text,
        ms_text,
        last,
    ]


def export_clients_xlsx(
    L: str = "en",
    q: str = "",
    group: str = "",
    account: str = "all",
) -> tuple[BytesIO, str]:
    clients = _filter_clients(load_clients(), q=q, group=group, account=account)
    wb = Workbook()
    ws = wb.active
    ws.title = extra_t(L, "export_sheet")[:31]

    headers = _export_headers(L)
    header_fill = PatternFill("solid", fgColor="007AFF")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, client in enumerate(clients, start=2):
        for col_idx, value in enumerate(_export_row(client, L), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in range(2, len(clients) + 2):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = date.today().isoformat()
    filename = f"CheckItNow-clients-{stamp}.xlsx"
    return buf, filename
