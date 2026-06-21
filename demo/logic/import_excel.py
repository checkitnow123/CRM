"""Import clients from Excel (.xlsx) — template or exported workbook."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from logic.clients import add_client, clear_all_clients, load_clients, normalize_client_numbers, save_clients
from logic.constants import LANG_CODES, MODE_MILESTONE, MODE_SINGLE
from logic.dates import today
from logic.extra_i18n import extra_t
from logic.i18n import mode_label, t

IMPORT_FIELDS = (
    "client_name",
    "group",
    "client_numbers",
    "contact_name",
    "client_email",
    "service_label",
    "mode",
    "end_date",
    "start_date",
)


def _norm_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("/", " ").split())


def _field_aliases() -> dict[str, set[str]]:
    aliases: dict[str, set[str]] = {field: set() for field in IMPORT_FIELDS}
    header_keys = {
        "client_name": lambda L: t(L, "col_client"),
        "group": lambda L: extra_t(L, "col_group"),
        "client_numbers": lambda L: extra_t(L, "col_client_no"),
        "contact_name": lambda L: extra_t(L, "col_contact"),
        "client_email": lambda L: extra_t(L, "col_email"),
        "service_label": lambda L: extra_t(L, "export_col_service"),
        "mode": lambda L: extra_t(L, "col_mode"),
        "end_date": lambda L: t(L, "col_end"),
        "start_date": lambda L: extra_t(L, "export_col_start"),
    }
    for lang in LANG_CODES:
        for field, getter in header_keys.items():
            try:
                title = getter(lang)
                if isinstance(title, tuple):
                    for part in title:
                        aliases[field].add(_norm_header(part))
                else:
                    aliases[field].add(_norm_header(title))
            except KeyError:
                pass
        aliases["end_date"].add(_norm_header(extra_t(lang, "col_end")))
        aliases["end_date"].add(_norm_header(t(lang, "col_end")))
    aliases["client_name"].update({"client name", "name", "company"})
    aliases["group"].update({"group"})
    aliases["client_numbers"].update({"client numbers", "numbers", "client no"})
    aliases["contact_name"].update({"contact", "contact name"})
    aliases["client_email"].update({"email", "e-mail"})
    aliases["service_label"].update({"service", "service label", "plan"})
    aliases["mode"].update({"type", "mode", "client type"})
    aliases["end_date"].update({"end", "expiry", "expiry date", "due date"})
    aliases["start_date"].update({"start", "start date"})
    return aliases


def _map_headers(headers: list[Any]) -> dict[str, int]:
    aliases = _field_aliases()
    lookup: dict[str, str] = {}
    for field, names in aliases.items():
        for name in names:
            lookup[name] = field
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        field = lookup.get(_norm_header(header))
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _parse_date(value: Any) -> date | None:
    if value is None or value == "" or value == "—" or value == "-":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or text in {"—", "-"}:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_mode(value: Any, L: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return MODE_SINGLE
    milestone_labels = {
        _norm_header(mode_label(lang, MODE_MILESTONE))
        for lang in LANG_CODES
    }
    milestone_labels.update({"milestone", "milestones", "里程碑"})
    if _norm_header(text) in milestone_labels or "milestone" in text or "里程碑" in text:
        return MODE_MILESTONE
    return MODE_SINGLE


def _parse_numbers(value: Any) -> list[str]:
    if value is None or value == "" or value == "—":
        return []
    if isinstance(value, (list, tuple)):
        return normalize_client_numbers([str(v) for v in value])
    return normalize_client_numbers(str(value).replace(";", ",").split(","))


def _cell(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_import_workbook(data: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Empty workbook"]
    mapping = _map_headers(list(rows[0]))
    if "client_name" not in mapping:
        return [], ["Missing client name column — use the import template or exported headers"]
    parsed: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = str(_cell(row, mapping, "client_name") or "").strip()
        if not name:
            continue
        mode_raw = _cell(row, mapping, "mode")
        end_raw = _cell(row, mapping, "end_date")
        end_date = _parse_date(end_raw)
        if end_date is None:
            errors.append(f"Row {row_num}: missing or invalid end date for {name!r}")
            continue
        parsed.append({
            "client_name": name,
            "group": str(_cell(row, mapping, "group") or "").strip(),
            "client_numbers": _parse_numbers(_cell(row, mapping, "client_numbers")),
            "contact_name": str(_cell(row, mapping, "contact_name") or "").strip(),
            "client_email": str(_cell(row, mapping, "client_email") or "").strip(),
            "service_label": str(_cell(row, mapping, "service_label") or "").strip(),
            "mode": _parse_mode(mode_raw, "en"),
            "end_date": end_date,
            "start_date": _parse_date(_cell(row, mapping, "start_date")),
        })
    return parsed, errors


def import_clients_from_rows(
    L: str,
    rows: list[dict[str, Any]],
    *,
    mode: str = "append",
) -> dict[str, Any]:
    if mode not in {"append", "replace"}:
        raise ValueError("Import mode must be append or replace.")
    if mode == "replace":
        clear_all_clients()
        clients: list[dict[str, Any]] = []
    else:
        clients = load_clients()

    imported = 0
    skipped = 0
    row_errors: list[str] = []
    for idx, row in enumerate(rows, start=1):
        name = row.get("client_name") or ""
        client_mode = row.get("mode") or MODE_SINGLE
        end_date = row.get("end_date")
        if not isinstance(end_date, date):
            skipped += 1
            row_errors.append(f"Row {idx}: invalid end date for {name!r}")
            continue
        if client_mode == MODE_MILESTONE:
            skipped += 1
            row_errors.append(
                f"Row {idx}: {name!r} — milestone clients need schedules; import as Single first, then set schedule in app"
            )
            continue
        try:
            entry = add_client(
                L,
                MODE_SINGLE,
                name,
                row.get("contact_name") or "",
                row.get("client_email") or "",
                row.get("service_label") or "",
                end_date,
                group=row.get("group") or "",
                client_numbers=row.get("client_numbers") or [],
            )
            if isinstance(row.get("start_date"), date):
                entry["start_date"] = row["start_date"]
            clients.append(entry)
            imported += 1
        except Exception as exc:
            skipped += 1
            row_errors.append(f"Row {idx}: {name!r} — {exc}")
    if imported:
        save_clients(clients)
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": row_errors,
        "total_clients": len(load_clients()) if imported else len(clients),
    }


def import_template_xlsx(L: str = "en") -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = extra_t(L, "import_sheet")[:31]
    headers = [
        t(L, "col_client"),
        extra_t(L, "col_group"),
        extra_t(L, "col_client_no"),
        extra_t(L, "col_contact"),
        extra_t(L, "col_email"),
        extra_t(L, "export_col_service"),
        extra_t(L, "col_mode"),
        t(L, "col_end"),
        extra_t(L, "export_col_start"),
    ]
    ws.append(headers)
    ws.append([
        "Example Co Ltd",
        "VIP",
        "1001, 1002",
        "Alex",
        "alex@example.com",
        "Annual plan",
        mode_label(L, MODE_SINGLE),
        today().isoformat(),
        today().isoformat(),
    ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "CheckItNow-import-template.xlsx"
